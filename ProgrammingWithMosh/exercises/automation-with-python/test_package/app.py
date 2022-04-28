import openpyxl as xl                               #this import allows us to open and make changes to excel workbooks
from openpyxl.chart import BarChart, Reference      #for creating a bar chart of the updated spreadsheet

def excel_example():
    # How to access excel workbook
    wb = xl.load_workbook('transactions.xlsx')

    # How to access an excel sheet
    sheet = wb['Sheet1']

    # How to access cells from a particular excel sheet
    # cell = sheet['a1'] 
    # cell sheet.cell(1, 1)          # A second approach

    cell = sheet['a1']

    # This shows the maximum count of rows what data occupies in the excel sheet
    # print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9              #this holds the value of the corrected price of column C
        corrected_price_cell = sheet.cell(row, 4)       #this assigns the value of corrected_price to column D
        corrected_price_cell.value = corrected_price    #this updates the spreadsheet with the value of correct_price populating column D

    # to create a bar chart, we use the reference class. the constructor is assigned the sheet as the main arguement, with
    # the minimum row defined at row 2 and the maximum row at the last row containing data
    # the min and max columns are defined to prevent from selecting any other data filled column if it were added
    values = Reference(sheet, 
                min_row=2, 
                max_row=sheet.max_row,
                min_col=4,
                max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save('transactions2.xlsx')                       #this saves the changes made to either the spreadsheet itself or another by assignment